#!/usr/bin/env python3
"""Fetch SimplifyJobs Summer 2026 internships and output an Excel spreadsheet sorted by region."""

import argparse
import json
import os
import re
from datetime import datetime, timedelta
from html.parser import HTMLParser

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

README_URL = (
    "https://raw.githubusercontent.com/SimplifyJobs/"
    "Summer2026-Internships/dev/README.md"
)
TRACKING_FILE = "simplify_tracking.json"

WESTERN_CA_PROVINCES = {"BC", "AB", "SK", "MB"}
EASTERN_CA_PROVINCES = {"ON", "QC", "NB", "NS", "PE", "NL", "NT", "NU", "YT"}
WESTERN_US_STATES = {
    "WA", "OR", "CA", "NV", "ID", "MT", "WY", "CO", "UT",
    "AZ", "NM", "HI", "AK",
}


# ---------------------------------------------------------------------------
# HTML table parser
# ---------------------------------------------------------------------------

class TableParser(HTMLParser):
    """Extract rows from all <table> elements in the document."""

    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._in_table = False
        self._in_row = False
        self._in_cell = False
        self._in_header = False
        self._current_row: list[str] = []
        self._current_cell = ""
        self._skip_row = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._in_table = True
        elif tag == "tr" and self._in_table:
            self._in_row = True
            self._current_row = []
            self._skip_row = False
        elif tag == "th" and self._in_row:
            self._in_header = True
            self._skip_row = True
            self._in_cell = True
            self._current_cell = ""
        elif tag == "td" and self._in_row:
            self._in_cell = True
            self._current_cell = ""
        elif tag == "a" and self._in_cell:
            for name, val in attrs:
                if name == "href":
                    self._current_cell += f"<a href=\"{val}\">"
        elif tag == "br" and self._in_cell:
            self._current_cell += "<br>"

    def handle_endtag(self, tag):
        if tag == "table":
            self._in_table = False
        elif tag == "tr" and self._in_row:
            self._in_row = False
            if not self._skip_row and self._current_row:
                self.rows.append(self._current_row)
        elif tag in ("td", "th") and self._in_cell:
            self._in_cell = False
            self._current_row.append(self._current_cell.strip())
        elif tag == "a" and self._in_cell:
            self._current_cell += "</a>"

    def handle_data(self, data):
        if self._in_cell:
            self._current_cell += data


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def extract_text(html_fragment: str) -> str:
    """Strip HTML tags, returning plain text."""
    return re.sub(r"<[^>]+>", "", html_fragment).strip()


def extract_first_link(html_fragment: str) -> str:
    """Return the href of the first <a> tag, stripping UTM params."""
    match = re.search(r'<a href="([^"]+)"', html_fragment)
    if not match:
        return ""
    url = match.group(1)
    url = re.sub(r"\?utm_source=.*", "", url)
    return url


def age_to_date(age_str: str) -> str:
    """Convert '3d' style age string to a date string like 'Apr 01, 2026'."""
    m = re.match(r"(\d+)d", age_str.strip())
    if m:
        days = int(m.group(1))
        dt = datetime.now() - timedelta(days=days)
        return dt.strftime("%b %d, %Y")
    return ""


def parse_date(date_str: str) -> datetime:
    """Parse 'Apr 01, 2026' style dates for sorting."""
    try:
        return datetime.strptime(date_str, "%b %d, %Y")
    except ValueError:
        return datetime(1970, 1, 1)


# ---------------------------------------------------------------------------
# Location expansion and classification
# ---------------------------------------------------------------------------

def expand_locations(location_html: str) -> list[str]:
    """Split a location cell (which may contain <br> or <details>) into individual locations."""
    # Strip <details>/<summary> wrapper if present
    text = re.sub(r"</?details>", "", location_html)
    text = re.sub(r"<summary>.*?</summary>", "", text)
    # Replace <br> with newline before stripping tags
    text = re.sub(r"<br\s*/?>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)  # strip remaining tags
    parts = [loc.strip() for loc in text.split("\n") if loc.strip()]
    return parts if parts else [text.strip()]


def classify_location(location_html: str) -> str:
    """Return one of: 'Western Canada', 'Eastern Canada', 'Western US', 'Other'."""
    locations = expand_locations(location_html)
    loc_text = " ".join(locations).strip()

    # Check for Canada
    if "canada" in loc_text.lower() or re.search(
        r"\b(" + "|".join(WESTERN_CA_PROVINCES | EASTERN_CA_PROVINCES) + r")\b,?\s*Canada",
        loc_text,
    ):
        # Check Western Canada
        for loc in locations:
            for prov in WESTERN_CA_PROVINCES:
                if re.search(rf"\b{prov}\b", loc):
                    return "Western Canada"
            if re.search(r"remote\s+in\s+canada", loc, re.IGNORECASE):
                return "Western Canada"
        # Check Eastern Canada
        for loc in locations:
            for prov in EASTERN_CA_PROVINCES:
                if re.search(rf"\b{prov}\b", loc):
                    return "Eastern Canada"
        # Generic Canada reference
        if "canada" in loc_text.lower():
            return "Western Canada"

    # Check for US states
    for loc in locations:
        if re.search(r"remote\s+in\s+usa", loc, re.IGNORECASE):
            return "Other"
        # Match "City, ST" pattern (US style, no country suffix)
        state_match = re.search(r",\s*([A-Z]{2})\s*$", loc.strip())
        if state_match:
            state = state_match.group(1)
            if state in WESTERN_US_STATES:
                return "Western US"
            return "Other"

    return "Other"


# ---------------------------------------------------------------------------
# Fetch and parse
# ---------------------------------------------------------------------------

def fetch_and_parse(url: str) -> list[dict]:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()

    parser = TableParser()
    parser.feed(resp.text)

    rows = []
    prev_company = ""

    for cols in parser.rows:
        if len(cols) < 5:
            continue

        company_html = cols[0]
        role = extract_text(cols[1])
        location_html = cols[2]
        apply_html = cols[3]
        age_str = extract_text(cols[4])

        company = extract_text(company_html)
        # Strip emoji prefixes
        company = re.sub(r"^[\U0001F525\U0001F6C2\U0001F1FA\U0001F1F8\s]+", "", company).strip()

        # Handle ↳ continuation rows
        if "↳" in company:
            company = prev_company
        else:
            prev_company = company

        apply_url = extract_first_link(apply_html)
        date_posted = age_to_date(age_str)
        location_text = "; ".join(expand_locations(location_html))
        region = classify_location(location_html)

        rows.append({
            "company": company,
            "role": role,
            "location": location_text,
            "apply_url": apply_url,
            "date_posted": date_posted,
            "region": region,
        })

    return rows


# ---------------------------------------------------------------------------
# Change tracking
# ---------------------------------------------------------------------------

def make_key(row: dict) -> str:
    """Create a stable identifier for a listing."""
    return f"{row['company']}|{row['role']}|{row['location']}"


def load_tracking(path: str) -> dict:
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def save_tracking(path: str, rows: list[dict], run_date: str) -> None:
    data = {
        "last_run": run_date,
        "listings": {make_key(r): {
            "company": r["company"],
            "role": r["role"],
            "location": r["location"],
            "apply_url": r["apply_url"],
            "date_posted": r["date_posted"],
            "region": r["region"],
            "first_seen": r.get("first_seen", run_date),
        } for r in rows},
    }
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def apply_tracking(rows: list[dict], tracking_path: str) -> list[dict]:
    """Compare current rows against previous snapshot. Returns rows with 'status' field added.

    Status values:
      'New'    — listing was not in the previous snapshot
      'Closed' — listing was in previous snapshot but is gone now
      ''       — listing existed before and is still open
    """
    today = datetime.now().strftime("%Y-%m-%d")
    prev = load_tracking(tracking_path)
    prev_listings = prev.get("listings", {})
    last_run = prev.get("last_run", "")

    current_keys = set()
    for row in rows:
        key = make_key(row)
        current_keys.add(key)
        if key in prev_listings:
            row["status"] = ""
            row["first_seen"] = prev_listings[key].get("first_seen", today)
        else:
            row["status"] = "New"
            row["first_seen"] = today

    # Find closed listings (were in previous snapshot, not in current)
    closed = []
    for key, prev_row in prev_listings.items():
        if key not in current_keys:
            prev_row["status"] = f"Closed (since {last_run})" if last_run else "Closed"
            prev_row["apply_url"] = ""
            closed.append(prev_row)

    all_rows = rows + closed
    save_tracking(tracking_path, rows, today)
    return all_rows


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

REGION_ORDER = ["Western Canada", "Eastern Canada", "Western US", "Other"]

REGION_COLORS = {
    "Western Canada": "2F5496",
    "Eastern Canada": "C00000",
    "Western US": "548235",
    "Other": "7F6000",
}


def write_xlsx(rows: list[dict], output: str) -> None:
    wb = Workbook()
    # Remove default sheet — we'll create one per region
    wb.remove(wb.active)

    headers = ["Status", "Company", "Role", "Location", "Apply Link", "Date Posted"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    body_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    wrap_alignment = Alignment(vertical="center", wrap_text=True)
    new_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    closed_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    closed_font = Font(name="Calibri", size=10, color="999999")

    # Group rows by region
    by_region: dict[str, list[dict]] = {r: [] for r in REGION_ORDER}
    for row in rows:
        by_region.setdefault(row["region"], by_region["Other"]).append(row)

    for region in REGION_ORDER:
        region_rows = by_region[region]
        if not region_rows:
            continue

        # Sort: New first, then open (blank status), then closed. Within each group, newest date first.
        def sort_key(r):
            status = r.get("status", "")
            if status == "New":
                priority = 0
            elif status.startswith("Closed"):
                priority = 2
            else:
                priority = 1
            return (priority, -parse_date(r["date_posted"]).timestamp())

        region_rows.sort(key=sort_key)

        ws = wb.create_sheet(title=region)
        header_fill = PatternFill(
            start_color=REGION_COLORS[region],
            end_color=REGION_COLORS[region],
            fill_type="solid",
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(region_rows, 2):
            status = row.get("status", "")
            values = [
                status, row["company"], row["role"], row["location"],
                row["apply_url"], row["date_posted"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Hyperlink for apply column
            if row["apply_url"]:
                link_cell = ws.cell(row=row_idx, column=5)
                link_cell.font = link_font
                link_cell.hyperlink = row["apply_url"]

            # Row highlighting
            if status == "New":
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = new_fill
            elif status.startswith("Closed"):
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.fill = closed_fill
                    c.font = closed_font

        col_widths = [22, 25, 50, 30, 55, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(region_rows) + 1}"

    wb.save(output)


def main():
    parser = argparse.ArgumentParser(
        description="Fetch SimplifyJobs internships sorted by region"
    )
    parser.add_argument(
        "-o", "--output", default="simplify_internships.xlsx",
        help="Output .xlsx file path (default: simplify_internships.xlsx)",
    )
    parser.add_argument(
        "--tracking-file", default=TRACKING_FILE,
        help=f"Path to tracking JSON file (default: {TRACKING_FILE})",
    )
    args = parser.parse_args()

    print("Fetching internship data from SimplifyJobs...")
    all_rows = fetch_and_parse(README_URL)
    print(f"  Found {len(all_rows)} active listings")

    all_rows = apply_tracking(all_rows, args.tracking_file)

    new_count = sum(1 for r in all_rows if r.get("status") == "New")
    closed_count = sum(1 for r in all_rows if r.get("status", "").startswith("Closed"))
    print(f"  {new_count} new, {closed_count} newly closed")

    for region in REGION_ORDER:
        count = sum(1 for r in all_rows if r["region"] == region)
        print(f"  {region}: {count}")

    write_xlsx(all_rows, args.output)
    print(f"Saved to {args.output}")


if __name__ == "__main__":
    main()
