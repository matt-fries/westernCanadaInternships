#!/usr/bin/env python3
"""Fetch Canadian Tech Internships 2026 and filter for Western Canada listings."""

import argparse
import re
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

README_URL = (
    "https://raw.githubusercontent.com/negarprh/"
    "Canadian-Tech-Internships-2026/main/README.md"
)

WESTERN_PROVINCES = {"AB", "BC", "SK", "MB"}
WESTERN_CITIES = {
    "calgary", "edmonton", "vancouver", "victoria", "burnaby",
    "kelowna", "richmond", "winnipeg", "saskatoon", "regina",
}
REMOTE_PATTERNS = re.compile(r"remote[,\s]*canada|remote\s+in\s+canada", re.IGNORECASE)


def is_western_canada(location: str) -> bool:
    loc_lower = location.lower()
    if REMOTE_PATTERNS.search(loc_lower):
        return True
    for city in WESTERN_CITIES:
        if city in loc_lower:
            return True
    for prov in WESTERN_PROVINCES:
        if re.search(rf"\b{prov}\b", location):
            return True
    return False


def parse_apply_cell(cell: str) -> tuple[str, str]:
    """Return (status, url) from the Apply column."""
    if "closed" in cell.lower():
        return "Closed", ""
    # Match nested Markdown image link: [![...](badge-url)](actual-url)
    nested = re.search(r"\[!\[.*?\]\(.*?\)\]\((https?://[^)]+)\)", cell)
    if nested:
        return "Open", nested.group(1)
    # Fallback: plain Markdown link [text](url)
    match = re.search(r"\]\((https?://[^)]+)\)", cell)
    if match:
        return "Open", match.group(1)
    return "Open", ""


def fetch_and_parse(url: str) -> list[dict]:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()

    rows = []
    in_table = False
    prev_company = ""

    for line in resp.text.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            if in_table:
                break  # table ended
            continue

        # Skip header separator row
        if re.match(r"\|[\s\-:]+\|", stripped):
            in_table = True
            continue

        # Detect header row (first row with |)
        if not in_table:
            in_table = True
            continue

        cols = [c.strip() for c in stripped.split("|")]
        # split on | gives empty strings at start/end
        cols = [c for c in cols if c or cols.index(c) not in (0, len(cols) - 1)]
        # Re-split properly
        cols = [c.strip() for c in stripped.strip("|").split("|")]

        if len(cols) < 5:
            continue

        company = cols[0].strip()
        role = cols[1].strip()
        location = cols[2].strip()
        apply_raw = cols[3].strip()
        date_posted = cols[4].strip()

        # Handle ↳ symbol
        if "↳" in company:
            company = prev_company
        else:
            prev_company = company

        status, apply_url = parse_apply_cell(apply_raw)

        rows.append({
            "company": company,
            "role": role,
            "location": location,
            "status": status,
            "apply_url": apply_url,
            "date_posted": date_posted,
        })

    return rows


def parse_date(date_str: str) -> datetime:
    """Parse 'Mar 31, 2026' style dates for sorting."""
    try:
        return datetime.strptime(date_str, "%b %d, %Y")
    except ValueError:
        return datetime(1970, 1, 1)


def write_xlsx(rows: list[dict], output: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Western Canada Internships"

    headers = ["Company", "Role", "Location", "Status", "Apply Link", "Date Posted"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Sort: open first, then by date descending
    rows.sort(key=lambda r: (r["status"] != "Open", parse_date(r["date_posted"])), reverse=False)
    # Secondary sort: open on top, then newest first within each group
    rows.sort(key=lambda r: (0 if r["status"] == "Open" else 1, parse_date(r["date_posted"])), reverse=False)
    # Actually: open first, then newest date first
    rows.sort(key=lambda r: (0 if r["status"] == "Open" else 1, -parse_date(r["date_posted"]).timestamp()))

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    body_font = Font(name="Calibri", size=10)
    wrap_alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        values = [
            row["company"], row["role"], row["location"],
            row["status"], row["apply_url"], row["date_posted"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = wrap_alignment

        # Apply link as plain URL
        if row["apply_url"]:
            link_cell = ws.cell(row=row_idx, column=5)
            link_cell.value = row["apply_url"]
            link_cell.font = link_font

        # Highlight open rows green
        if row["status"] == "Open":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = green_fill

    # Column widths
    col_widths = [25, 50, 25, 10, 50, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(output)


def main():
    parser = argparse.ArgumentParser(
        description="Fetch and filter Western Canada tech internships"
    )
    parser.add_argument(
        "-o", "--output", default="western_canada_internships.xlsx",
        help="Output .xlsx file path (default: western_canada_internships.xlsx)",
    )
    parser.add_argument(
        "--open-only", action="store_true",
        help="Exclude closed listings",
    )
    args = parser.parse_args()

    print(f"Fetching internship data from GitHub...")
    all_rows = fetch_and_parse(README_URL)
    print(f"  Found {len(all_rows)} total listings")

    western = [r for r in all_rows if is_western_canada(r["location"])]
    print(f"  {len(western)} match Western Canada filter")

    if args.open_only:
        western = [r for r in western if r["status"] == "Open"]
        print(f"  {len(western)} open listings after filtering")

    write_xlsx(western, args.output)
    print(f"Saved to {args.output}")


if __name__ == "__main__":
    main()
